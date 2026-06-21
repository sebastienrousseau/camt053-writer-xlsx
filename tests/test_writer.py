# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Tests for the camt053-writer-xlsx writer."""

from __future__ import annotations

from pathlib import Path

import pytest
from camt053.models import (
    Account,
    Balance,
    Entry,
    ParsedDocument,
    Statement,
    TransactionDetails,
)
from openpyxl import load_workbook

from camt053_writer_xlsx import __version__, write_xlsx


def _document_with_one_of_everything() -> ParsedDocument:
    """Build a small ParsedDocument exercising every column path."""
    return ParsedDocument(
        message_type="camt.053.001.08",
        msg_id="MSG-1",
        creation_date_time="2026-06-21T10:00:00",
        statements=[
            Statement(
                id="STMT-1",
                electronic_seq_nb="42",
                creation_date_time="2026-06-21T10:00:00",
                account=Account(
                    iban="DE89370400440532013000",
                    other_id=None,
                    currency="EUR",
                    owner_name="ACME GmbH",
                    servicer_bic="COBADEFFXXX",
                ),
                balances=[
                    Balance(
                        type_code="OPBD",
                        amount="1000.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                        date="2026-06-20",
                    ),
                    Balance(
                        type_code="CLBD",
                        amount="1500.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                        date="2026-06-21",
                    ),
                ],
                entries=[
                    Entry(
                        reference="NTRY-1",
                        amount="500.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                        status="BOOK",
                        booking_date="2026-06-21",
                        value_date="2026-06-21",
                        account_servicer_ref="REF-1",
                        reversal_indicator=False,
                        reason_code=None,
                        details=[
                            TransactionDetails(
                                end_to_end_id="E2E-1",
                                tx_id="TX-1",
                                instruction_id="INSTR-1",
                                reason_code=None,
                                counterparty_name="Customer X",
                                counterparty_account="FR1420041010050500013M02606",
                                additional_info="Invoice 123",
                            )
                        ],
                    ),
                    Entry(
                        reference="NTRY-2",
                        amount="100.00",
                        currency="EUR",
                        credit_debit_indicator="DBIT",
                        reversal_indicator=True,
                        reason_code="AC04",
                        details=[],
                    ),
                ],
            )
        ],
    )


def test_version_exposed() -> None:
    """The package exposes a non-empty semantic-style version string."""
    assert isinstance(__version__, str)
    assert __version__.count(".") >= 2


def test_write_xlsx_returns_path_and_writes_file(tmp_path: Path) -> None:
    """The function returns the output path and creates a non-empty file."""
    document = _document_with_one_of_everything()
    out = tmp_path / "out.xlsx"

    result = write_xlsx(document, out)

    assert result == out
    assert out.exists()
    assert out.stat().st_size > 0


def test_metadata_sheet_columns_and_row(tmp_path: Path) -> None:
    """The Metadata sheet has the canonical header and one row per statement."""
    document = _document_with_one_of_everything()
    out = tmp_path / "out.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    sheet = workbook["Metadata"]
    header = [c.value for c in sheet[1]]
    assert header == [
        "message_type",
        "msg_id",
        "creation_date_time",
        "statement_id",
        "electronic_seq_nb",
        "statement_creation_date_time",
        "account_iban",
        "account_other_id",
        "account_currency",
        "account_owner_name",
        "account_servicer_bic",
        "balance_count",
        "entry_count",
    ]
    row = [c.value for c in sheet[2]]
    assert row[0] == "camt.053.001.08"
    assert row[3] == "STMT-1"
    assert row[11] == 2  # balance_count
    assert row[12] == 2  # entry_count


def test_balances_sheet_populated(tmp_path: Path) -> None:
    """Every balance gets its own row."""
    document = _document_with_one_of_everything()
    out = tmp_path / "out.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    sheet = workbook["Balances"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    type_codes = [r[1] for r in rows]
    assert type_codes == ["OPBD", "CLBD"]


def test_entries_sheet_flattens_details(tmp_path: Path) -> None:
    """An entry with N details yields N rows; a detail-less entry yields one."""
    document = _document_with_one_of_everything()
    out = tmp_path / "out.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    sheet = workbook["Entries"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # NTRY-1 has 1 detail → 1 row; NTRY-2 has no details → 1 row.
    assert len(rows) == 2
    detail_e2e = [r[11] for r in rows]
    assert detail_e2e == ["E2E-1", None]


def test_reversals_sheet_only_returnable_entries(tmp_path: Path) -> None:
    """The Reversals sheet filters to entries with reversal flag or reason."""
    document = _document_with_one_of_everything()
    out = tmp_path / "out.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    sheet = workbook["Reversals"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # Only NTRY-2 (reversal_indicator=True, reason_code="AC04") qualifies.
    assert len(rows) == 1
    assert rows[0][1] == "NTRY-2"
    assert rows[0][10] == "AC04"


def test_entry_with_multiple_details_is_flattened(tmp_path: Path) -> None:
    """An entry with two details produces two rows that share entry columns."""
    document = ParsedDocument(
        message_type="camt.053.001.02",
        msg_id="MSG-2",
        creation_date_time="2026-06-21T10:00:00",
        statements=[
            Statement(
                id="STMT-X",
                account=Account(iban="GB29NWBK60161331926819"),
                balances=[],
                entries=[
                    Entry(
                        reference="NTRY-A",
                        amount="250.00",
                        currency="GBP",
                        credit_debit_indicator="CRDT",
                        details=[
                            TransactionDetails(end_to_end_id="A"),
                            TransactionDetails(end_to_end_id="B"),
                        ],
                    ),
                ],
            )
        ],
    )
    out = tmp_path / "multi-detail.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    rows = list(workbook["Entries"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    # Both rows carry the same entry-level reference.
    assert {r[1] for r in rows} == {"NTRY-A"}
    # And each one carries a distinct detail end-to-end ID.
    assert {r[11] for r in rows} == {"A", "B"}


def test_accepts_string_path(tmp_path: Path) -> None:
    """The path argument accepts a str as well as a Path."""
    document = _document_with_one_of_everything()
    out = tmp_path / "string-path.xlsx"

    result = write_xlsx(document, str(out))

    assert result == out


def test_empty_document_writes_only_headers(tmp_path: Path) -> None:
    """A document with zero statements still produces a valid workbook."""
    document = ParsedDocument(message_type="camt.053.001.08")
    out = tmp_path / "empty.xlsx"

    write_xlsx(document, out)

    workbook = load_workbook(out)
    for name in ("Metadata", "Balances", "Entries", "Reversals"):
        sheet = workbook[name]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows == []


def test_returns_resolved_path_type(tmp_path: Path) -> None:
    """Path argument round-trips so callers can chain on the return value."""
    document = ParsedDocument(message_type="camt.053.001.08")
    out = tmp_path / "type.xlsx"

    result = write_xlsx(document, out)

    assert isinstance(result, Path)


def test_unwritable_path_raises_oserror(tmp_path: Path) -> None:
    """An unwritable target surfaces an OSError unchanged."""
    document = ParsedDocument(message_type="camt.053.001.08")
    # A directory rather than a file path → openpyxl can't write to it.
    with pytest.raises(OSError):
        write_xlsx(document, tmp_path)
