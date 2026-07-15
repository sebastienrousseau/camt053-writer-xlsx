# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Load / stress tests for the camt053-writer-xlsx writer.

Marked ``perf`` and excluded from the default run and its coverage
gate (mirroring the ``camt053`` parent suite convention). Select them
explicitly with::

    pytest tests/test_stress.py -m perf --no-cov

Three scenarios:

* sustained concurrent writes — many threads serialising a
  representative statement to distinct temp files, asserting zero
  errors and a generous p95 latency bound;
* a large workbook — a statement with several thousand entries within
  a generous wall-clock bound and a bounded ``tracemalloc`` peak;
* a soak loop — repeated writes asserting bounded memory growth.

Bounds are deliberately loose: they exist to catch pathological
regressions (deadlocks, quadratic blow-ups, unbounded caches), not to
benchmark. Keep the whole module under ~60 s on developer hardware.
"""

from __future__ import annotations

import gc
import time
import tracemalloc
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pytest
from camt053.models import (
    Account,
    Balance,
    Entry,
    ParsedDocument,
    Statement,
    TransactionDetails,
)
from openpyxl import load_workbook

from camt053_writer_xlsx import write_xlsx

pytestmark = pytest.mark.perf

# --- Tunables (generous on purpose; see module docstring) -----------

CONCURRENT_WORKERS = 16
CONCURRENT_ITERATIONS = 160
CONCURRENT_P95_SECONDS = 5.0

LARGE_ENTRY_COUNT = 4000
LARGE_WALL_CLOCK_SECONDS = 30.0
LARGE_PEAK_BYTES = 512 * 1024 * 1024  # 512 MiB

SOAK_ITERATIONS = 30
SOAK_WARMUP_ITERATIONS = 5
SOAK_GROWTH_BYTES = 16 * 1024 * 1024  # 16 MiB


def _entry(index: int) -> Entry:
    """Build one booked entry with a single transaction detail."""
    return Entry(
        reference=f"NTRY-{index}",
        amount=f"{100 + index % 900}.{index % 100:02d}",
        currency="EUR",
        credit_debit_indicator="CRDT" if index % 2 == 0 else "DBIT",
        status="BOOK",
        booking_date="2026-06-21",
        value_date="2026-06-21",
        account_servicer_ref=f"REF-{index}",
        reversal_indicator=index % 25 == 0,
        reason_code="AC04" if index % 25 == 0 else None,
        details=[
            TransactionDetails(
                end_to_end_id=f"E2E-{index}",
                tx_id=f"TX-{index}",
                instruction_id=f"INSTR-{index}",
                reason_code=None,
                counterparty_name=f"Counterparty {index}",
                counterparty_account="FR1420041010050500013M02606",
                additional_info=f"Invoice {index}",
            )
        ],
    )


def _document(entry_count: int) -> ParsedDocument:
    """Build a representative single-statement parsed document."""
    return ParsedDocument(
        message_type="camt.053.001.08",
        msg_id=f"MSG-STRESS-{entry_count}",
        creation_date_time="2026-06-21T10:00:00",
        statements=[
            Statement(
                id=f"STMT-STRESS-{entry_count}",
                electronic_seq_nb="1",
                creation_date_time="2026-06-21T10:00:00",
                account=Account(
                    iban="DE89370400440532013000",
                    other_id=None,
                    currency="EUR",
                    owner_name="ACME GmbH",
                    servicer_bic="COBADEFFXXX",
                ),
                balances=[
                    Balance(
                        type_code="OPBD",
                        amount="1000.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                        date="2026-06-20",
                    ),
                    Balance(
                        type_code="CLBD",
                        amount="1500.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                        date="2026-06-21",
                    ),
                ],
                entries=[_entry(i) for i in range(entry_count)],
            )
        ],
    )


def _p95(samples: list[float]) -> float:
    """Return the 95th-percentile value of ``samples``."""
    ordered = sorted(samples)
    index = min(len(ordered) - 1, int(0.95 * len(ordered)))
    return ordered[index]


def test_sustained_concurrent_writes(tmp_path: Path) -> None:
    """Many threads writing distinct workbooks: zero errors, sane p95."""
    document = _document(entry_count=50)
    latencies: list[float] = []
    errors: list[BaseException] = []

    def _write_one(index: int) -> float:
        """Write one workbook to a unique path; return elapsed seconds."""
        target = tmp_path / f"concurrent-{index}.xlsx"
        started = time.perf_counter()
        result = write_xlsx(document, target)
        elapsed = time.perf_counter() - started
        assert result == target
        assert target.stat().st_size > 0
        return elapsed

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as pool:
        futures = [
            pool.submit(_write_one, i) for i in range(CONCURRENT_ITERATIONS)
        ]
        for future in as_completed(futures):
            exc = future.exception()
            if exc is not None:
                errors.append(exc)
            else:
                latencies.append(future.result())

    assert errors == []
    assert len(latencies) == CONCURRENT_ITERATIONS
    assert _p95(latencies) < CONCURRENT_P95_SECONDS

    # Spot-check one output is a loadable workbook with the full sheet
    # set, so "zero errors" means real files, not just no exceptions.
    workbook = load_workbook(tmp_path / "concurrent-0.xlsx")
    assert workbook.sheetnames == [
        "Metadata",
        "Balances",
        "Entries",
        "Reversals",
    ]


def test_large_workbook_within_bounds(tmp_path: Path) -> None:
    """Several thousand entries stay within wall-clock + memory bounds."""
    document = _document(entry_count=LARGE_ENTRY_COUNT)
    target = tmp_path / "large.xlsx"

    gc.collect()
    tracemalloc.start()
    started = time.perf_counter()
    write_xlsx(document, target)
    elapsed = time.perf_counter() - started
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    assert elapsed < LARGE_WALL_CLOCK_SECONDS
    assert peak < LARGE_PEAK_BYTES

    workbook = load_workbook(target, read_only=True)
    entries_sheet = workbook["Entries"]
    row_count = sum(1 for _ in entries_sheet.iter_rows(min_row=2))
    workbook.close()
    assert row_count == LARGE_ENTRY_COUNT


def test_soak_repeated_writes_bounded_memory(tmp_path: Path) -> None:
    """Repeated writes must not leak: traced memory growth is bounded."""
    document = _document(entry_count=50)
    target = tmp_path / "soak.xlsx"

    tracemalloc.start()
    try:
        for _ in range(SOAK_WARMUP_ITERATIONS):
            write_xlsx(document, target)
        gc.collect()
        baseline, _ = tracemalloc.get_traced_memory()

        for _ in range(SOAK_ITERATIONS):
            write_xlsx(document, target)
        gc.collect()
        final, _ = tracemalloc.get_traced_memory()
    finally:
        tracemalloc.stop()

    growth = final - baseline
    assert growth < SOAK_GROWTH_BYTES, (
        f"traced memory grew by {growth} bytes over "
        f"{SOAK_ITERATIONS} iterations"
    )
