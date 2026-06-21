# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Build a ParsedDocument programmatically and inspect the Reversals sheet.

This example shows the most common auditor workflow: extract returnable
entries from a statement and surface them on their own sheet. Run with
``python examples/02_filter_reversals.py``.
"""

from pathlib import Path

from camt053.models import Account, Entry, ParsedDocument, Statement
from openpyxl import load_workbook

from camt053_writer_xlsx import write_xlsx


def main() -> None:
    """Write a document with one normal entry and one reversal."""
    document = ParsedDocument(
        message_type="camt.053.001.08",
        msg_id="MSG-DEMO-REV",
        creation_date_time="2026-06-21T10:00:00",
        statements=[
            Statement(
                id="STMT-REV",
                account=Account(iban="DE89370400440532013000", currency="EUR"),
                entries=[
                    Entry(
                        reference="NTRY-OK",
                        amount="500.00",
                        currency="EUR",
                        credit_debit_indicator="CRDT",
                    ),
                    Entry(
                        reference="NTRY-RET",
                        amount="100.00",
                        currency="EUR",
                        credit_debit_indicator="DBIT",
                        reversal_indicator=True,
                        reason_code="AC04",
                    ),
                ],
            )
        ],
    )
    out = Path("out_reversals.xlsx")
    write_xlsx(document, out)

    workbook = load_workbook(out)
    reversal_rows = list(
        workbook["Reversals"].iter_rows(min_row=2, values_only=True)
    )
    print(f"Wrote {out.resolve()}; reversal rows: {len(reversal_rows)}")
    for row in reversal_rows:
        print(f"  {row[1]} amount={row[2]} {row[3]} reason={row[10]}")


if __name__ == "__main__":
    main()
