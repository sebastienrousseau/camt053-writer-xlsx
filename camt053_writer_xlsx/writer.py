# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Multi-sheet Excel writer for parsed camt.053 documents.

The workbook layout is deliberately stable so downstream tooling
(reconciliation macros, audit pivot tables) can target columns by name
without parsing the document model.

Sheets
------
Metadata
    One row per statement with its identifying header fields.
Balances
    One row per balance reported across all statements.
Entries
    One row per booked entry across all statements. Multi-detail
    entries are flattened: an entry with N TransactionDetails yields N
    rows that share the entry-level columns.
Reversals
    A filter of Entries where ``reversal_indicator`` is true or any
    return reason code is present. Useful as the auditor's first view.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from camt053.models import (
    Balance,
    Entry,
    ParsedDocument,
    Statement,
    TransactionDetails,
)
from openpyxl import Workbook
from openpyxl.styles import Font

__all__ = ["write_xlsx"]


_METADATA_COLUMNS = [
    "message_type",
    "msg_id",
    "creation_date_time",
    "statement_id",
    "electronic_seq_nb",
    "statement_creation_date_time",
    "account_iban",
    "account_other_id",
    "account_currency",
    "account_owner_name",
    "account_servicer_bic",
    "balance_count",
    "entry_count",
]

_BALANCE_COLUMNS = [
    "statement_id",
    "type_code",
    "amount",
    "currency",
    "credit_debit_indicator",
    "date",
]

_ENTRY_COLUMNS = [
    "statement_id",
    "entry_reference",
    "amount",
    "currency",
    "credit_debit_indicator",
    "status",
    "booking_date",
    "value_date",
    "account_servicer_ref",
    "reversal_indicator",
    "entry_reason_code",
    "detail_end_to_end_id",
    "detail_tx_id",
    "detail_instruction_id",
    "detail_reason_code",
    "detail_counterparty_name",
    "detail_counterparty_account",
    "detail_additional_info",
]


def write_xlsx(
    document: ParsedDocument,
    path: str | Path,
) -> Path:
    """Serialise ``document`` to a multi-sheet ``.xlsx`` workbook.

    Args:
        document: The parsed camt.053 document to serialise.
        path: The output ``.xlsx`` file path. Parent directories must
            already exist; the file is overwritten if it already exists.

    Returns:
        The resolved :class:`~pathlib.Path` of the written workbook.

    Raises:
        OSError: If the file cannot be written (e.g. permission denied).
    """
    output = Path(path)
    workbook = Workbook()

    metadata_sheet = workbook.active
    metadata_sheet.title = "Metadata"
    _write_metadata(metadata_sheet, document)

    balances_sheet = workbook.create_sheet("Balances")
    _write_balances(balances_sheet, document)

    entries_sheet = workbook.create_sheet("Entries")
    _write_entries(entries_sheet, document)

    reversals_sheet = workbook.create_sheet("Reversals")
    _write_reversals(reversals_sheet, document)

    workbook.save(output)
    return output


def _write_header(sheet: Any, columns: list[str]) -> None:
    """Write a bold header row of ``columns`` to ``sheet`` at row 1."""
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _write_metadata(sheet: Any, document: ParsedDocument) -> None:
    """Populate the Metadata sheet, one row per statement."""
    _write_header(sheet, _METADATA_COLUMNS)
    for statement in document.statements:
        sheet.append(_metadata_row(document, statement))


def _metadata_row(
    document: ParsedDocument,
    statement: Statement,
) -> list[Any]:
    """Return the metadata row for ``statement``."""
    account = statement.account
    return [
        document.message_type,
        document.msg_id,
        document.creation_date_time,
        statement.id,
        statement.electronic_seq_nb,
        statement.creation_date_time,
        account.iban,
        account.other_id,
        account.currency,
        account.owner_name,
        account.servicer_bic,
        len(statement.balances),
        len(statement.entries),
    ]


def _write_balances(sheet: Any, document: ParsedDocument) -> None:
    """Populate the Balances sheet, one row per balance per statement."""
    _write_header(sheet, _BALANCE_COLUMNS)
    for statement in document.statements:
        for balance in statement.balances:
            sheet.append(_balance_row(statement, balance))


def _balance_row(statement: Statement, balance: Balance) -> list[Any]:
    """Return the balance row for ``balance`` under ``statement``."""
    return [
        statement.id,
        balance.type_code,
        balance.amount,
        balance.currency,
        balance.credit_debit_indicator,
        balance.date,
    ]


def _write_entries(sheet: Any, document: ParsedDocument) -> None:
    """Populate the Entries sheet, one row per (entry, detail) pair."""
    _write_header(sheet, _ENTRY_COLUMNS)
    for statement in document.statements:
        for entry in statement.entries:
            for row in _entry_rows(statement, entry):
                sheet.append(row)


def _entry_rows(
    statement: Statement,
    entry: Entry,
) -> Iterable[list[Any]]:
    """Yield one row per detail; one bare-entry row if no details exist.

    Multi-detail entries are flattened so the workbook stays a flat
    rectangular shape (which is what Excel pivots and accounting macros
    expect).
    """
    if not entry.details:
        yield _entry_row(statement, entry, None)
        return
    for detail in entry.details:
        yield _entry_row(statement, entry, detail)


def _entry_row(
    statement: Statement,
    entry: Entry,
    detail: TransactionDetails | None,
) -> list[Any]:
    """Return one Entries row for ``entry`` (optionally with one ``detail``)."""
    return [
        statement.id,
        entry.reference,
        entry.amount,
        entry.currency,
        entry.credit_debit_indicator,
        entry.status,
        entry.booking_date,
        entry.value_date,
        entry.account_servicer_ref,
        entry.reversal_indicator,
        entry.reason_code,
        detail.end_to_end_id if detail else None,
        detail.tx_id if detail else None,
        detail.instruction_id if detail else None,
        detail.reason_code if detail else None,
        detail.counterparty_name if detail else None,
        detail.counterparty_account if detail else None,
        detail.additional_info if detail else None,
    ]


def _write_reversals(sheet: Any, document: ParsedDocument) -> None:
    """Populate the Reversals sheet with returnable / reversed entries."""
    _write_header(sheet, _ENTRY_COLUMNS)
    for statement in document.statements:
        for entry in statement.entries:
            if not (entry.reversal_indicator or entry.is_returnable()):
                continue
            for row in _entry_rows(statement, entry):
                sheet.append(row)
